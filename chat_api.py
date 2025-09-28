from flask import Blueprint, request, jsonify, render_template_string
import json
import os
from datetime import datetime
import re
import openpyxl
import io
import traceback

# Crear blueprint para el chat
chat_bp = Blueprint('chat', __name__)

# Archivo de datos independiente del chat
CHAT_DATOS_FILE = 'chat_datos.json'

def cargar_datos_chat():
    """Cargar datos específicos del chat (independientes del boletín)"""
    try:
        if not os.path.exists(CHAT_DATOS_FILE):
            return None
        
        with open(CHAT_DATOS_FILE, 'r', encoding='utf-8') as f:
            datos = json.load(f)
        return datos
    except Exception as e:
        print(f"Error cargando datos del chat: {e}")
        return None

def leer_excel_chat_y_convertir(archivo_excel):
    """Convertir Excel del chat a formato JSON - VERSIÓN CORREGIDA"""
    try:
        print(f"🔍 Chat: Procesando archivo: {archivo_excel.filename}")
        
        archivo_excel.seek(0)
        file_content = archivo_excel.read()
        excel_buffer = io.BytesIO(file_content)
        
        print("📖 Chat: Cargando workbook...")
        workbook = openpyxl.load_workbook(excel_buffer, data_only=True)
        
        datos_chat = {
            'fecha_carga': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'tribunales': {}
        }
        
        print(f"📊 Chat: Hojas encontradas: {workbook.sheetnames}")
        
        # Procesar cada hoja como tribunal independiente
        for sheet_name in workbook.sheetnames:
            print(f"📄 Chat: Procesando hoja: {sheet_name}")
            sheet = workbook[sheet_name]
            
            # Leer headers (primera fila) - MANERA SEGURA
            headers = []
            for cell in sheet[1]:
                if cell.value is None:
                    headers.append('')
                else:
                    try:
                        headers.append(str(cell.value).strip())
                    except:
                        headers.append('')
            
            print(f"📋 Chat: Headers para {sheet_name}: {headers}")
            
            # Leer datos (desde fila 2) - MANERA SEGURA
            sheet_data = []
            row_count = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_count += 1
                # Verificar si la fila tiene datos
                if any(cell is not None for cell in row):
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            # Convertir valor de manera segura
                            if value is None:
                                row_dict[headers[i]] = ''
                            elif isinstance(value, datetime):
                                row_dict[headers[i]] = value.strftime('%Y-%m-%d')
                            else:
                                try:
                                    row_dict[headers[i]] = str(value)
                                except:
                                    row_dict[headers[i]] = ''
                    
                    if row_dict:  # Solo agregar si tiene datos
                        sheet_data.append(row_dict)
            
            if sheet_data:
                datos_chat['tribunales'][sheet_name] = sheet_data
                print(f"✅ Chat: {sheet_name}: {len(sheet_data)} registros")
        
        print(f"🎉 Chat: Procesamiento completado. Total hojas: {len(datos_chat['tribunales'])}")
        return datos_chat
        
    except Exception as e:
        print(f"❌ Error detallado en chat: {str(e)}")
        print(f"📝 Stack trace: {traceback.format_exc()}")
        raise Exception(f"Error procesando Excel del chat: {str(e)}")

def parse_query_basico(query_text):
    """Parser básico para extraer filtros de consultas en lenguaje natural"""
    query_lower = query_text.lower()
    filtros = {}
    
    # Detectar expediente
    exp_pattern = r'(?:expediente|exp\.?|tf)\s*[-\s]*(\d+[-/]\w*)'
    exp_match = re.search(exp_pattern, query_lower)
    if exp_match:
        filtros['expediente'] = exp_match.group(1)
    
    # Detectar año
    year_pattern = r'\b(20\d{2})\b'
    year_match = re.search(year_pattern, query_text)
    if year_match:
        filtros['año'] = int(year_match.group(1))
    
    # Detectar sala
    sala_pattern = r'sala\s*([a-g]|[1-7])'
    sala_match = re.search(sala_pattern, query_lower)
    if sala_match:
        filtros['sala'] = sala_match.group(1).upper()
    
    # Detectar tribunal
    if 'tfn' in query_lower or 'tribunal fiscal' in query_lower:
        filtros['tribunal'] = 'TFN'
    elif 'cncaf' in query_lower or 'cámara' in query_lower:
        filtros['tribunal'] = 'CNCAF'
    elif 'csjn' in query_lower or 'corte suprema' in query_lower:
        filtros['tribunal'] = 'CSJN'
    
    # Detectar temas comunes
    temas = {
        'prescripción': ['prescripcion', 'prescripc'],
        'honorarios': ['honorario'],
        'infracciones': ['infraccion', 'infrac'],
        'nulidad': ['nulidad'],
        'apelación': ['apelacion', 'recurso']
    }
    
    for tema, keywords in temas.items():
        if any(keyword in query_lower for keyword in keywords):
            filtros['tema'] = tema
            break
    
    return filtros

def filtrar_datos_chat(datos_chat, filtros):
    """Aplicar filtros a los datos del chat"""
    if not datos_chat or 'tribunales' not in datos_chat:
        return []
    
    resultados = []
    
    # Buscar en todas las hojas/tribunales
    for hoja_name, registros in datos_chat['tribunales'].items():
        for item in registros:
            if coincide_filtros_chat(item, filtros, hoja_name):
                item['_fuente'] = hoja_name
                resultados.append(item)
    
    return resultados

def coincide_filtros_chat(item, filtros, hoja_name):
    """Verificar si un item del chat coincide con los filtros"""
    
    # Filtro por expediente - buscar en cualquier campo que contenga "expediente"
    if 'expediente' in filtros:
        expediente_encontrado = False
        for key, value in item.items():
            if 'expediente' in key.lower() and value:
                if filtros['expediente'].lower() in str(value).lower():
                    expediente_encontrado = True
                    break
        if not expediente_encontrado:
            return False
    
    # Filtro por sala - buscar en cualquier campo que contenga "sala"
    if 'sala' in filtros:
        sala_encontrada = False
        for key, value in item.items():
            if 'sala' in key.lower() and value:
                if str(value).upper() == filtros['sala']:
                    sala_encontrada = True
                    break
        if not sala_encontrada:
            return False
    
    # Filtro por tema - buscar en campos de tema, carátula, resuelve
    if 'tema' in filtros:
        tema_encontrado = False
        campos_tema = [key for key in item.keys() if any(x in key.lower() for x in ['tema', 'caratula', 'resuelve'])]
        for campo in campos_tema:
            if item.get(campo) and filtros['tema'].lower() in str(item[campo]).lower():
                tema_encontrado = True
                break
        if not tema_encontrado:
            return False
    
    # Filtro por año - buscar en fechas o en nombre de hoja
    if 'año' in filtros:
        año_encontrado = False
        
        # Buscar en campos de fecha
        campos_fecha = [key for key in item.keys() if 'fecha' in key.lower()]
        for campo in campos_fecha:
            if item.get(campo) and str(filtros['año']) in str(item[campo]):
                año_encontrado = True
                break
        
        # Buscar en nombre de hoja
        if not año_encontrado and str(filtros['año']) in hoja_name:
            año_encontrado = True
        
        # Si no hay fechas específicas, asumir que puede coincidir
        if not año_encontrado and not campos_fecha:
            año_encontrado = True
    
    # Filtro por tribunal - buscar en nombre de hoja o campos
    if 'tribunal' in filtros:
        tribunal_encontrado = False
        
        # Buscar en nombre de hoja
        if filtros['tribunal'].lower() in hoja_name.lower():
            tribunal_encontrado = True
        
        # Buscar en campos que contengan tribunal
        if not tribunal_encontrado:
            campos_tribunal = [key for key in item.keys() if 'tribunal' in key.lower()]
            for campo in campos_tribunal:
                if item.get(campo) and filtros['tribunal'].lower() in str(item[campo]).lower():
                    tribunal_encontrado = True
                    break
        
        if not tribunal_encontrado:
            return False
    
    return True

def generar_respuesta_chat(query, filtros, resultados):
    """Generar respuesta conversacional para el chat"""
    total = len(resultados)
    
    if total == 0:
        return {
            "mensaje": f"No encontré resultados para '{query}' en la base de datos del chat.",
            "sugerencias": [
                "Verifica que los datos estén cargados correctamente",
                "Prueba términos más generales",
                "Revisa la ortografía de los filtros"
            ]
        }
    
    # Construir mensaje principal
    mensaje_parts = [f"Encontré {total} resultado{'s' if total != 1 else ''}"]
    
    if filtros.get('tribunal'):
        mensaje_parts.append(f"en {filtros['tribunal']}")
    
    if filtros.get('tema'):
        mensaje_parts.append(f"sobre {filtros['tema']}")
    
    if filtros.get('sala'):
        mensaje_parts.append(f"de la sala {filtros['sala']}")
    
    if filtros.get('año'):
        mensaje_parts.append(f"del año {filtros['año']}")
    
    mensaje = " ".join(mensaje_parts) + "."
    
    # Análisis de fuentes
    fuentes = {}
    for item in resultados:
        fuente = item.get('_fuente', 'Desconocido')
        fuentes[fuente] = fuentes.get(fuente, 0) + 1
    
    analisis = []
    if len(fuentes) > 1:
        fuente_info = []
        for fuente, count in fuentes.items():
            fuente_info.append(f"{count} en {fuente}")
        analisis.append("Distribución: " + ", ".join(fuente_info))
    
    return {
        "mensaje": mensaje,
        "analisis": analisis,
        "fuentes": list(fuentes.keys())
    }

# ENDPOINTS DEL CHAT

@chat_bp.route('/test', methods=['GET'])
def test_chat():
    """Endpoint de prueba para verificar funcionamiento del chat"""
    datos = cargar_datos_chat()
    
    total_registros = 0
    tribunales_info = {}
    
    if datos and 'tribunales' in datos:
        for tribunal, registros in datos['tribunales'].items():
            count = len(registros)
            tribunales_info[tribunal] = count
            total_registros += count
    
    return jsonify({
        "status": "ok",
        "message": "Chat API funcionando correctamente",
        "timestamp": datetime.now().isoformat(),
        "data_available": datos is not None,
        "total_registros": total_registros,
        "tribunales_disponibles": tribunales_info,
        "ultima_carga": datos.get('fecha_carga') if datos else None
    })

@chat_bp.route('/upload', methods=['POST'])
def subir_datos_chat():
    """Endpoint para cargar datos específicos del chat"""
    try:
        if 'archivo' not in request.files:
            return jsonify({'error': 'No se encontró archivo'}), 400
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        if not archivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Solo se permiten archivos Excel (.xlsx, .xls)'}), 400
        
        print(f"Procesando archivo del chat: {archivo.filename}")
        
        # Procesar Excel específico del chat
        datos_chat = leer_excel_chat_y_convertir(archivo)
        
        # Guardar en archivo JSON del chat
        with open(CHAT_DATOS_FILE, 'w', encoding='utf-8') as f:
            json.dump(datos_chat, f, ensure_ascii=False, indent=2)
        
        print("Datos del chat guardados correctamente")
        
        # Calcular estadísticas
        total_registros = sum(len(registros) for registros in datos_chat['tribunales'].values())
        
        return jsonify({
            'mensaje': 'Datos del chat procesados exitosamente',
            'fecha_carga': datos_chat['fecha_carga'],
            'total_registros': total_registros,
            'tribunales_cargados': list(datos_chat['tribunales'].keys()),
            'detalle_por_tribunal': {k: len(v) for k, v in datos_chat['tribunales'].items()}
        })
        
    except Exception as e:
        print(f"Error en subir_datos_chat: {str(e)}")
        return jsonify({'error': str(e)}), 500

@chat_bp.route('/query', methods=['POST'])
def procesar_consulta_chat():
    """Endpoint principal para procesar consultas del chat"""
    try:
        data = request.get_json()
        
        if not data or 'query' not in data:
            return jsonify({
                "success": False,
                "error": "Se requiere el campo 'query' en el request"
            }), 400
        
        query = data['query'].strip()
        
        if not query:
            return jsonify({
                "success": False,
                "error": "La consulta no puede estar vacía"
            }), 400
        
        # Cargar datos del chat
        datos_chat = cargar_datos_chat()
        if not datos_chat:
            return jsonify({
                "success": False,
                "error": "No hay datos del chat disponibles. Carga un archivo Excel primero."
            }), 404
        
        # Procesar consulta
        filtros = parse_query_basico(query)
        resultados = filtrar_datos_chat(datos_chat, filtros)
        respuesta = generar_respuesta_chat(query, filtros, resultados)
        
        # Preparar respuesta
        response_data = {
            "success": True,
            "query": query,
            "filtros_detectados": filtros,
            "total_resultados": len(resultados),
            "respuesta": respuesta,
            "datos": resultados[:10] if len(resultados) <= 10 else resultados[:5],
            "hay_mas_resultados": len(resultados) > 10
        }
        
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Error procesando consulta del chat: {str(e)}"
        }), 500

@chat_bp.route('/status', methods=['GET'])
def status_chat():
    """Estado del sistema de chat independiente"""
    datos = cargar_datos_chat()
    
    status_info = {
        "chat_enabled": True,
        "data_last_update": datos.get('fecha_carga') if datos else None,
        "sistema": "independiente_del_boletin"
    }
    
    if datos and 'tribunales' in datos:
        status_info["tribunales_disponibles"] = {
            tribunal: len(registros) 
            for tribunal, registros in datos['tribunales'].items()
        }
        status_info["total_registros"] = sum(
            len(registros) for registros in datos['tribunales'].values()
        )
    else:
        status_info["tribunales_disponibles"] = {}
        status_info["total_registros"] = 0
    
    status_info["supported_queries"] = [
        "Búsqueda por expediente: 'expediente TF-12345'",
        "Filtro por tema: 'sentencias sobre prescripción'",
        "Filtro por sala: 'casos de la sala G'",
        "Filtro por año: 'sentencias de 2023'",
        "Combinaciones: 'casos de prescripción sala G 2023'"
    ]
    
    return jsonify(status_info)

@chat_bp.route('/admin', methods=['GET'])
def admin_chat():
    """Página de administración específica del chat"""
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Admin Chat - Sistema de Jurisprudencia</title>
        <meta charset="UTF-8">
        <style>
            body { font-family: Arial, sans-serif; max-width: 600px; margin: 50px auto; padding: 20px; }
            .upload-area { border: 2px dashed #007cba; padding: 40px; text-align: center; margin: 20px 0; border-radius: 10px; }
            button { background: #007cba; color: white; padding: 12px 24px; border: none; border-radius: 5px; cursor: pointer; font-size: 16px; }
            button:hover { background: #005a8b; }
            .status { margin: 20px 0; padding: 15px; border-radius: 5px; }
            .success { background: #d4edda; color: #155724; border: 1px solid #c3e6cb; }
            .error { background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb; }
            .info { background: #e2e3e5; color: #383d41; border: 1px solid #d6d8db; }
        </style>
    </head>
    <body>
        <h1>Panel de Administración - CHAT</h1>
        <h2>Sistema Independiente de Jurisprudencia</h2>
        
        <div class="upload-area">
            <h3>Cargar datos del Chat</h3>
            <p>Archivo Excel con datos históricos de jurisprudencia (independiente del boletín diario)</p>
            <input type="file" id="fileInput" accept=".xlsx,.xls" style="margin: 10px;">
            <br><br>
            <button onclick="subirArchivo()">Cargar Datos del Chat</button>
        </div>
        
        <div id="status"></div>
        
        <div style="margin-top: 30px;">
            <h3>Enlaces del Chat:</h3>
            <p><a href="/api/chat/test" target="_blank">Test del Chat</a></p>
            <p><a href="/api/chat/status" target="_blank">Estado del Chat</a></p>
            <p><strong>Endpoint principal:</strong><br>
            <code>POST /api/chat/query</code></p>
        </div>
        
        <script>
            function subirArchivo() {
                const fileInput = document.getElementById('fileInput');
                const statusDiv = document.getElementById('status');
                
                if (!fileInput.files[0]) {
                    statusDiv.innerHTML = '<div class="error">Por favor selecciona un archivo</div>';
                    return;
                }
                
                const formData = new FormData();
                formData.append('archivo', fileInput.files[0]);
                
                statusDiv.innerHTML = '<div class="info">Procesando datos del chat...</div>';
                
                fetch('/api/chat/upload', {
                    method: 'POST',
                    body: formData
                })
                .then(response => response.json())
                .then(data => {
                    if (data.error) {
                        statusDiv.innerHTML = `<div class="error">Error: ${data.error}</div>`;
                    } else {
                        statusDiv.innerHTML = `<div class="success">
                            ✓ Datos del chat cargados exitosamente<br>
                            Fecha: ${data.fecha_carga}<br>
                            Total registros: ${data.total_registros}<br>
                            Tribunales: ${data.tribunales_cargados.join(', ')}
                        </div>`;
                        fileInput.value = '';
                    }
                })
                .catch(error => {
                    statusDiv.innerHTML = `<div class="error">Error de conexión: ${error}</div>`;
                });
            }
        </script>
    </body>
    </html>
    ''')