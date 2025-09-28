from flask import Flask, request, jsonify, render_template_string
from flask_cors import CORS
import openpyxl
import json
import os
from datetime import datetime
import io
import traceback

app = Flask(__name__)
CORS(app)

# Archivo donde se guardan los datos DEL BOLETIN (independiente del chat)
DATOS_FILE = 'datos.json'

def leer_excel_y_convertir(archivo_excel):
    """Convierte el Excel a formato JSON usando openpyxl - VERSIÓN CORREGIDA"""
    try:
        print(f"🔍 Procesando archivo: {archivo_excel.filename}")
        
        # En lugar de pasar el archivo directamente, lo leemos en memoria
        archivo_excel.seek(0)
        file_content = archivo_excel.read()
        
        # Crear un objeto BytesIO para simular un archivo
        excel_buffer = io.BytesIO(file_content)
        
        # Cargar el workbook desde el buffer
        workbook = openpyxl.load_workbook(excel_buffer, data_only=True)
        
        datos = {
            'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'tfn': [],
            'tfn_cncaf': [],
            'tfn_cncaf_csjn': []
        }
        
        print(f"📊 Hojas encontradas: {workbook.sheetnames}")
        
        # Procesar cada hoja
        sheet_mapping = {
            'TFN': 'tfn',
            'TFN_CNCAF': 'tfn_cncaf', 
            'TFN_CNCAF_CSJN': 'tfn_cncaf_csjn'
        }
        
        for sheet_name, data_key in sheet_mapping.items():
            if sheet_name in workbook.sheetnames:
                print(f"📄 Procesando hoja: {sheet_name}")
                sheet = workbook[sheet_name]
                
                # Leer headers (primera fila) - CON DEBUG DETALLADO
                headers = []
                print(f"🔍 DEBUG - Leyendo headers de {sheet_name}:")
                for i, cell in enumerate(sheet[1]):
                    if cell.value is None:
                        headers.append('')
                        print(f"   Celda {i}: None -> ''")
                    else:
                        # Convertir a string de manera segura CON DEBUG
                        try:
                            original_value = cell.value
                            converted_value = str(original_value).strip()
                            headers.append(converted_value)
                            print(f"   Celda {i}: '{original_value}' -> '{converted_value}'")
                        except Exception as conv_error:
                            headers.append('')
                            print(f"   Celda {i}: ERROR en conversión: {conv_error}")
                
                print(f"📋 Headers finales para {sheet_name}: {headers}")
                
                # NORMALIZAR NOMBRES DE COLUMNAS - CORRECCIÓN CRÍTICA
                headers_normalizados = []
                for header in headers:
                    if header:
                        # Corregir errores comunes de tipeo
                        header_normalizado = header.replace('Garatula_TFM', 'Caratula_TFN')
                        header_normalizado = header_normalizado.replace('Competencia_TFM', 'Competencia_TFN')
                        header_normalizado = header_normalizado.replace('Expediente_TFM', 'Expediente_TFN')
                        header_normalizado = header_normalizado.replace('Sala_TFM', 'Sala_TFN')
                        header_normalizado = header_normalizado.replace('Vocalia_TFM', 'Vocalia_TFN')
                        header_normalizado = header_normalizado.replace('Resuelve_TFM', 'Resuelve_TFN')
                        header_normalizado = header_normalizado.replace('Tema_TFM', 'Tema_TFN')
                        headers_normalizados.append(header_normalizado)
                    else:
                        headers_normalizados.append('')
                
                print(f"🔧 Headers normalizados: {headers_normalizados}")
                
                # Leer datos (desde fila 2 en adelante) - MANERA SEGURA
                sheet_data = []
                row_count = 0
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_count += 1
                    # Verificar si la fila tiene datos (no todos None)
                    if any(cell is not None for cell in row):
                        row_dict = {}
                        for i, value in enumerate(row):
                            if i < len(headers_normalizados) and headers_normalizados[i]:
                                # Convertir valor de manera segura
                                if value is None:
                                    row_dict[headers_normalizados[i]] = ''
                                elif isinstance(value, datetime):
                                    row_dict[headers_normalizados[i]] = value.strftime('%Y-%m-%d %H:%M:%S')
                                else:
                                    try:
                                        row_dict[headers_normalizados[i]] = str(value)
                                    except:
                                        row_dict[headers_normalizados[i]] = ''
                        
                        if row_dict:  # Solo agregar si tiene datos
                            sheet_data.append(row_dict)
                
                datos[data_key] = sheet_data
                print(f"✅ {sheet_name}: {len(sheet_data)} registros procesados")
        
        print(f"🎉 Procesamiento completado: TFN={len(datos['tfn'])}, CNCAF={len(datos['tfn_cncaf'])}, CSJN={len(datos['tfn_cncaf_csjn'])}")
        return datos
        
    except Exception as e:
        print(f"❌ Error detallado en leer_excel_y_convertir: {str(e)}")
        print(f"📝 Stack trace: {traceback.format_exc()}")
        raise Exception(f"Error procesando Excel: {str(e)}")

@app.route('/')
def dashboard():
    """Sirve el dashboard principal desde index.html"""
    try:
        with open('index.html', 'r', encoding='utf-8') as f:
            return f.read()
    except FileNotFoundError:
        return "Error: index.html no encontrado. Asegúrate de que el archivo esté en el repositorio.", 404

@app.route('/api/status')
def api_status():
    """Endpoint para verificar el estado del backend"""
    return "Backend del Boletín de Trazabilidad funcionando correctamente"

@app.route('/admin')
def admin():
    """Página simple para subir archivos DEL BOLETIN"""
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Admin - Boletín de Trazabilidad</title>
        <meta charset="UTF-8">
        <style>
            body { font-family: Arial, sans-serif; max-width: 600px; margin: 50px auto; padding: 20px; }
            .upload-area { border: 2px dashed #ccc; padding: 40px; text-align: center; margin: 20px 0; border-radius: 10px; }
            button { background: #007cba; color: white; padding: 12px 24px; border: none; border-radius: 5px; cursor: pointer; font-size: 16px; }
            button:hover { background: #005a8b; }
            .status { margin: 20px 0; padding: 15px; border-radius: 5px; }
            .success { background: #d4edda; color: #155724; border: 1px solid #c3e6cb; }
            .error { background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb; }
            .info { background: #e2e3e5; color: #383d41; border: 1px solid #d6d8db; }
            .separator { margin: 40px 0; border-top: 2px solid #007cba; padding-top: 20px; }
        </style>
    </head>
    <body>
        <h1>Panel de Administración</h1>
        <h2>Boletín de Trazabilidad de Sentencias (DIARIO)</h2>
        
        <div class="upload-area">
            <h3>Subir archivo Excel - BOLETIN DIARIO</h3>
            <p>Selecciona el archivo Excel con las 3 hojas: TFN, TFN_CNCAF, TFN_CNCAF_CSJN</p>
            <input type="file" id="fileInput" accept=".xlsx,.xls" style="margin: 10px;">
            <br><br>
            <button onclick="subirArchivo()">Actualizar Boletín</button>
        </div>
        
        <div id="status"></div>
        
        <div class="separator">
            <h2>Sistema de Chat (INDEPENDIENTE)</h2>
            <p><a href="/api/chat/admin" style="background: #28a745; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
                Ir al Panel del Chat
            </a></p>
            <small>El chat maneja datos históricos independientes del boletín diario</small>
        </div>
        
        <div style="margin-top: 30px;">
            <h3>Enlaces útiles del Boletín:</h3>
            <p><a href="/">Ver Dashboard Principal</a></p>
            <p><a href="/api/datos" target="_blank">Ver datos JSON del boletín</a></p>
            <p><a href="/api/test" target="_blank">Test de funcionalidad</a></p>
            <p><strong>URL del API del Boletín:</strong><br>
            <code>https://tfndata.onrender.com/api/datos</code></p>
        </div>
        
        <script>
            function subirArchivo() {
                const fileInput = document.getElementById('fileInput');
                const statusDiv = document.getElementById('status');
                
                if (!fileInput.files[0]) {
                    statusDiv.innerHTML = '<div class="error">Por favor selecciona un archivo</div>';
                    return;
                }
                
                const formData = new FormData();
                formData.append('archivo', fileInput.files[0]);
                
                statusDiv.innerHTML = '<div class="info">Procesando archivo...</div>';
                
                fetch('/api/subir', {
                    method: 'POST',
                    body: formData
                })
                .then(response => response.json())
                .then(data => {
                    if (data.error) {
                        statusDiv.innerHTML = `<div class="error">Error: ${data.error}</div>`;
                    } else {
                        statusDiv.innerHTML = `<div class="success">
                            ✓ Archivo procesado exitosamente<br>
                            Fecha: ${data.fecha_actualizacion}<br>
                            TFN: ${data.total_tfn} registros<br>
                            TFN-CNCAF: ${data.total_tfn_cncaf} registros<br>
                            TFN-CNCAF-CSJN: ${data.total_tfn_cncaf_csjn} registros
                        </div>`;
                        fileInput.value = '';
                    }
                })
                .catch(error => {
                    statusDiv.innerHTML = `<div class="error">Error de conexión: ${error}</div>`;
                });
            }
        </script>
    </body>
    </html>
    ''')

@app.route('/api/subir', methods=['POST'])
def subir_archivo():
    """Endpoint para subir y procesar el Excel DEL BOLETIN"""
    try:
        if 'archivo' not in request.files:
            return jsonify({'error': 'No se encontró archivo'}), 400
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        if not archivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Solo se permiten archivos Excel (.xlsx, .xls)'}), 400
        
        print(f"Procesando archivo: {archivo.filename}")  # Log para Render
        
        # Procesar Excel
        datos = leer_excel_y_convertir(archivo)
        
        print(f"Datos procesados - TFN: {len(datos['tfn'])}, TFN_CNCAF: {len(datos['tfn_cncaf'])}, TFN_CNCAF_CSJN: {len(datos['tfn_cncaf_csjn'])}")
        
        # Guardar en archivo JSON
        with open(DATOS_FILE, 'w', encoding='utf-8') as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)
        
        print("Archivo JSON guardado correctamente")
        
        # Respuesta con estadísticas
        return jsonify({
            'mensaje': 'Archivo procesado exitosamente',
            'fecha_actualizacion': datos['fecha_actualizacion'],
            'total_tfn': len(datos['tfn']),
            'total_tfn_cncaf': len(datos['tfn_cncaf']),
            'total_tfn_cncaf_csjn': len(datos['tfn_cncaf_csjn'])
        })
        
    except Exception as e:
        print(f"Error en subir_archivo: {str(e)}")  # Log para debugging
        return jsonify({'error': str(e)}), 500

@app.route('/api/datos')
def obtener_datos():
    """Endpoint que devuelve los datos para el frontend DEL BOLETIN"""
    try:
        if not os.path.exists(DATOS_FILE):
            return jsonify({'error': 'No hay datos disponibles. Sube un archivo Excel primero.'}), 404
        
        with open(DATOS_FILE, 'r', encoding='utf-8') as f:
            datos = json.load(f)
        
        return jsonify(datos)
        
    except Exception as e:
        print(f"Error en obtener_datos: {str(e)}")
        return jsonify({'error': f'Error cargando datos: {str(e)}'}), 500

@app.route('/api/test')
def test():
    """Endpoint para probar que todo funciona"""
    test_info = {
        'status': 'OK',
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'datos_file_exists': os.path.exists(DATOS_FILE),
        'current_directory': os.getcwd(),
        'files_in_directory': os.listdir('.') if os.path.exists('.') else [],
        'environment': 'RENDER' if 'RENDER' in os.environ else 'LOCAL'
    }
    
    if os.path.exists(DATOS_FILE):
        file_size = os.path.getsize(DATOS_FILE)
        test_info['datos_file_size_bytes'] = file_size
        test_info['datos_file_size_mb'] = round(file_size / (1024 * 1024), 2)
        
        try:
            with open(DATOS_FILE, 'r', encoding='utf-8') as f:
                datos = json.load(f)
            test_info['data_summary'] = {
                'fecha_actualizacion': datos.get('fecha_actualizacion'),
                'tfn_records': len(datos.get('tfn', [])),
                'tfn_cncaf_records': len(datos.get('tfn_cncaf', [])),
                'tfn_cncaf_csjn_records': len(datos.get('tfn_cncaf_csjn', []))
            }
        except Exception as e:
            test_info['error_reading_data'] = str(e)
    
    return jsonify(test_info)

# INTEGRACION DEL CHAT - Solo estas 3 líneas agregadas
try:
    from chat_api import chat_bp
    app.register_blueprint(chat_bp, url_prefix='/api/chat')
    print("✓ Chat API integrado correctamente")
except ImportError as e:
    print(f"⚠ Chat API no disponible: {e}")

if __name__ == '__main__':
    # Configuración para Render
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') == 'development'
    
    app.run(debug=debug_mode, host='0.0.0.0', port=port)